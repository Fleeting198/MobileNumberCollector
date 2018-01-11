from bs4 import BeautifulSoup
from Utils import *
import openpyxl
from uniformLogger import logger
import os, shutil, traceback, re

"""
从手机取证系统输出的html报告里提取包含传文件的聊天记录，结果写进xlsx
要根据具体报告格式写代码，现在写了readChats1和readChats2两种
"""

QQs={"A":['12345678'],
     "B":['87654321']}



class ChatItem:
    def __init__(self, userFrom, userTo, cDate, cTime, content, MD5=""):
        self.userFrom = userFrom
        self.userTo = userTo
        self.cDate = cDate
        self.cTime = cTime
        self.content = content
        self.file_path = ""
        self.cnt_mobile = ""
        self.MD5 = MD5


class MobileEvidenceHtmlChatFinder:
    targetExts = ['xls', 'zip', 'rar']

    def __init__(self, root_evidence, path_opt_sheet, root_search, root_exception):
        """
        :param root_evidence: 报告文件所在目录
        :param path_opt_sheet: 输出报告文件的路径
        :param root_search: 去此目录中搜索文件计算手机号数
        :param root_exception: 手机号数计算失败的文件复制到本目录中
        """
        self.root_evidence = root_evidence
        self.path_opt_sheet = path_opt_sheet
        self.root_search = root_search
        self.root_exception = root_exception
        if not os.path.isdir(root_search):
            logger.exception("root_search not exist.")
            raise FileNotFoundError("root_search not exist.")
        if not os.path.isdir(root_exception):
            os.makedirs(root_exception)

        self.chats = []

    def handle_cnt_mobile(self, chat):
        # 调用cnt_mobile_content，根据返回值给chat设置文件路径、手机号数、备注等信息
        result, str_pathes = self.cnt_mobile_content(chat)
        chat.file_path = str_pathes
        if result:
            chat.cnt_mobile = result

    def cnt_mobile_content(self, chat):
        """计数chat中聊天内容中的文件包含的手机号数"""
        if not self.is_target_chat_ext(chat):
            return None, ""
        file_name = chat.content.split('\\')[-1]
        possible_dirs = search_file(file_name, self.root_search)
        str_paths = ';'.join(possible_dirs)

        if not possible_dirs:
            result = -1
            logger.info("Failed locating file: {}.".format(file_name))
        else:
            result = self.cnt_from_multi_path(possible_dirs)
        return result, str_paths

    def cnt_from_multi_path(self, possibleDirs):
        """ 若去重后所有可能路径的结果相同，返回这个手机号数结果，否则返回None"""
        counts = [self.wrapper_count_mobile(p) for p in possibleDirs]
        counts = list(set(counts))  # 去重
        if len(counts) == 1:
            return counts[0]
        return None

    def wrapper_count_mobile(self, file):
        """ 计数file中的手机号数，若出错或为零则复制到root_exception下"""
        try:
            num = len(count_unique_mobile(file))
            if num == 0:
                raise Exception
            else:
                return num
        except Exception:
            try:
                fileName = file.split('\\')[-1]
                tarPath = self.root_exception + '\\' + fileName
                shutil.copy(file, tarPath)
                logger.info("Copy file to exception: {}.".format(tarPath))
                return None
            except Exception:
                # 在复制时出错
                traceback.print_exc()

    def readChats1(self, html):
        rePattern = r"发送人：(.+)\s+日期：(\d{4}-\d{2}-\d{2})\s(\d{2}:\d{2}:\d{2})<br>\s+内容：(.+)<br><br>"
        reComplier = re.compile(rePattern)

        matches = reComplier.findall(html)
        for match in matches:
            cfrom = match[0].strip()
            cdate = match[1]
            ctime = match[2]
            ccontent = match[3]
            chat = ChatItem(cfrom, "", cdate, ctime, ccontent)
            if self.is_target_chat_ext(chat):
                self.chats.append(chat)

    def readChats2(self, html):
        soup = BeautifulSoup(html, 'lxml')
        try:
            trs = soup.body.table.find_all('tr', recursive=False)
            for tr in trs:
                trclass = tr.get('class', None)
                if trclass and trclass[0] == 'tablehead':
                    continue

                tds = tr.find_all('td')

                cfrom = tds[1].get_text(strip=True)
                cto = tds[2].get_text(strip=True)
                tmp = tds[3].get_text(strip=True)

                cdate = tmp[10:]
                ctime = tmp[:10]

                tmp = tds[4].find_all('td')
                ccontent = tmp[1].get_text(strip=True)
                cmd5 = tmp[3].get_text(strip=True)

                chat = ChatItem(cfrom, cto, cdate, ctime, ccontent, cmd5)
                if self.is_target_chat_qq(chat, '人名'):
                    self.chats.append(chat)
        except AttributeError:
            pass
        except IndexError:
            pass

    def process(self):
        for file in files(self.root_evidence):
            if ext_in(file, ['html', 'htm']):
                # self.readChats1(self.readHtml(file))
                self.readChats2(self.read_html(file))

        for c in self.chats:
            self.handle_cnt_mobile(c)
        self.writeToExcel()

    def writeToExcel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        titles = ['发送人', '接收人', '日期', '时间', '内容', '文件路径', '手机号数', 'MD5']
        for i, title in enumerate(titles):
            ws.cell(row=1, column=i + 1, value=title)

        for i, chat in enumerate(self.chats):
            rowID = i + 2
            ws.cell(row=rowID, column=1, value=chat.userFrom)
            ws.cell(row=rowID, column=2, value=chat.userTo)
            ws.cell(row=rowID, column=3, value=chat.cDate)
            ws.cell(row=rowID, column=4, value=chat.cTime)
            ws.cell(row=rowID, column=5, value=chat.content)
            ws.cell(row=rowID, column=6, value=chat.file_path)
            ws.cell(row=rowID, column=7, value=chat.cnt_mobile)
            if chat.MD5:
                ws.cell(row=rowID, column=6, value=chat.MD5)

        wb.save(self.path_opt_sheet)

    def read_html(self, file):
        print(file)
        with open(file, 'rb') as f:
            html = f.read()
            return html

    def is_target_chat_ext(self, chat):
        for ext in MobileEvidenceHtmlChatFinder.targetExts:
            if ext in str(chat.content).lower():
                return True
        return False

    def is_target_chat_qq(self, chat, exclude_person):
        for k, vgroup in QQs.items():
            if k == exclude_person: continue
            for v in vgroup:
                if v in str(chat.userTo) or v in str(chat.userFrom):
                    return True

        return False
