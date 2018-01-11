from Utils import *
import openpyxl
import os


class ReportFileGatherer:
    """ 遍历报告excel文件，对每个报告，读取记录中的文件路径，从root_search中复制出来放到一个文件夹中，并计数去重后的手机号数，输出到控制台"""

    def __init__(self, root_src, root_search, report_file):
        self.root_src = root_src
        self.rootSearch = root_search
        self.reportFile = report_file
        self.results = []

    def recordFiles(self):
        for file in files(self.root_src):
            p1, ext = os.path.splitext(file)
            if p1.endswith("文件") and ext == ".xlsx":
                yield file

    def recordResults(self):
        with open(self.reportFile, 'w') as f:
            for record in self.results:
                f.write(record[0] + ": " + record[1] + '\n')

    def procedure(self):
        for file in self.recordFiles():
            print('\n')
            print(file)
            filesFolder = self.prepareFileFolder(file)
            self.copyFiles(file, filesFolder)
            result = self.countUniqueMobilesInFolder(filesFolder)

            print("文件夹去重手机号数" + str(result))
            self.results.append([file, str(result)])
        self.recordResults()

    def prepareFileFolder(self, file):
        p = file.split('.')[0]
        if not os.path.isdir(p):
            os.makedirs(p)
        return p

    def copyFiles(self, file, filesFolder):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        for i in range(2, ws.max_row):
            cellFilePath = ws.cell(row=i, column=6).value
            cellContent = ws.cell(row=i, column=5).value
            cellNumMobile = ws.cell(row=i, column=7).value

            if cellNumMobile and cellFilePath:
                cellFilePath = cellFilePath.split(";")

                p1, p2 = os.path.split(cellContent)
                for possibleDir in cellFilePath:
                    src = self.rootSearch + '\\' + possibleDir + '\\' + p2
                    des = filesFolder
                    import shutil
                    shutil.copy(src, des)
                    # copyCheckExist(src, des)

    def countUniqueMobilesInFolder(self, folder):
        mobiles = set()
        for file in files(folder):
            for item in count_unique_mobile(file):
                mobiles.add(item)

        return len(mobiles)
